from pathlib import Path
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

def process_workbook(names,prices,soorten,alcoholpercentages,volumes,webpages,images):
    wb = xl.load_workbook('Drinks.xlsx')
    sheet = wb['Blad1']

    for i in range(len(names)):
        sheet.cell(i + 2, 1).value = names[i]
        sheet.cell(i + 2, 2).value = prices[i]
        sheet.cell(i + 2, 3).value = soorten[i]
        sheet.cell(i + 2, 4).value = alcoholpercentages[i]
        sheet.cell(i + 2, 5).value = volumes[i]
        sheet.cell(i + 2, 7).value = 'https://www.gall.nl'+webpages[i]
        sheet.cell(i + 2, 8).value = images[i]
        print("opgeslagen")
    wb.save('Drinks.xlsx')


